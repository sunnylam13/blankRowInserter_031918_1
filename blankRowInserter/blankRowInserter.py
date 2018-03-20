# -*- coding: utf-8 -*-

#! python3

import openpyxl, sys

from openpyxl.styles import Font
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell

try:
	from openpyxl.cell import column_index_from_string,get_column_letter
except ImportError:
	from openpyxl.utils import column_index_from_string,get_column_letter

import logging
logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
# logging.disable(logging.CRITICAL)

# parse the command line

position_row_N = sys.argv[1]
position_row_N = 3 # for testing
logging.debug('The row where we start insertion is:  %s' % (position_row_N))

blank_row_num_M = sys.argv[2]
blank_row_num_M = 2 # for testing
logging.debug('The number of blank rows to insert is:  %s' % (blank_row_num_M))

sheet_to_process = sys.argv[3]
sheet_to_process = "../tests/updatedProduceSales.xlsx" # set it for testing purposes, disable in final product
logging.debug('The spreadsheet to process is:  %s' % (sheet_to_process))

# access the workbook

wb = openpyxl.load_workbook(sheet_to_process) # create the workbook
sheet = wb.active # switch to the active sheet, there should only be one

logging.debug('Testing to see that sheet loaded right, giving a value:  ')
logging.debug(sheet['A2'])
logging.debug(sheet['A2'].value)

# find out the max rows and max columns so we can set upper ends for loops

upper_row_max = sheet.max_row
logging.debug('The maximum number of rows in the sheet is %i' % (upper_row_max))

upper_col_max = sheet.max_column
logging.debug('The maximum number of columns in the sheet is %i' % (upper_col_max))

# copy all of the data up to the N row and its cells, insertion of M happens after N

values_uptoN_list = [] # to store the values from the spreadsheet
values_afterM_list = [] # to store the values from the spreadsheet, after M blank lines inserted

value_uptoN_dict = {}
values_afterM_dict = {}

# list version
# def row_analyzer(values_list,min_value,max_value):
	
# 	for rowValue in range(min_value,max_value): # +1 because we're not starting at 0
# 		# do not set the upper limit to upper_row_max+1 as normal, instead set it to n + 1 as we will be inserting the gap at that point
# 		# we still add + 1 because range() stops 1 point before position_row_N normally, we want it to stop exactly at position_row_N
		
# 		# within this specific rowValue we go through each colValue
# 		for colValue in range(1,upper_col_max+1): # +1 because we're not starting at 0
# 			# convert the colValue into a letter coordinate
# 			column_letter = get_column_letter(colValue)
# 			# combine the column coordinate and row coordinate
# 			cell_coordinate = column_letter + str(rowValue)
# 			# get the cell coordinate's value and push it into the values_list
# 			cell_value = sheet[cell_coordinate].value
			
# 			values_list.append(cell_value)
# 			# logging.debug('The value for %s has been stored in the values_list' % (cell_coordinate))
# 			# logging.debug('The value for %s' % (cell_coordinate) )
# 			# logging.debug(cell_value)

# dict version
def row_analyzer(values_dict,min_value,max_value,blank_rows_insert=0):
	
	for rowValue in range(min_value,max_value): # +1 because we're not starting at 0
		# do not set the upper limit to upper_row_max+1 as normal, instead set it to n + 1 as we will be inserting the gap at that point
		# we still add + 1 because range() stops 1 point before position_row_N normally, we want it to stop exactly at position_row_N
		
		# within this specific rowValue we go through each colValue
		for colValue in range(1,upper_col_max+1): # +1 because we're not starting at 0
			# convert the colValue into a letter coordinate
			column_letter = get_column_letter(colValue)
			# combine the column coordinate and row coordinate
			cell_coordinate = column_letter + str(rowValue + blank_rows_insert)
			# get the cell coordinate's value and push it into the values_list
			cell_value = sheet[cell_coordinate].value
			
			# values_list.append(cell_value)
			# logging.debug('The value for %s has been stored in the values_list' % (cell_coordinate))
			# logging.debug('The value for %s' % (cell_coordinate) )
			# logging.debug(cell_value)

			values_dict[cell_coordinate] = cell_value
			# logging.debug('The value for %s has been stored in the values_dict' % (cell_coordinate))
			# logging.debug('The value for %s' % (cell_coordinate) )
			# logging.debug(cell_value)

# def row_builder(values_list,min_value,max_value,worksheet):
	
# 	for rowValue in range(min_value,max_value): # +1 because we're not starting at 0
# 		# do not set the upper limit to upper_row_max+1 as normal, instead set it to n + 1 as we will be inserting the gap at that point
# 		# we still add + 1 because range() stops 1 point before position_row_N normally, we want it to stop exactly at position_row_N
		
# 		# within this specific rowValue we go through each colValue
# 		for colValue in range(1,upper_col_max+1): # +1 because we're not starting at 0
# 			# convert the colValue into a letter coordinate
# 			column_letter = get_column_letter(colValue)
# 			# combine the column coordinate and row coordinate
# 			cell_coordinate = column_letter + str(rowValue) # a string
# 			# go to the `values_list` and set the value of the cell to that value
# 			worksheet[cell_coordinate] = values_list[colValue]


# 			# # get the cell coordinate's value and push it into the values_list
# 			# cell_value = sheet[column_letter + str(rowValue)].value

# 			# values_list.append(cell_value)
# 			# # logging.debug('The value for %s has been stored in the values_list' % (column_letter + str(rowValue)))
# 			# # logging.debug('The value for %s' % (column_letter + str(rowValue)) )
# 			# # logging.debug(cell_value)

def row_builder(values_dict,workbook):

	sheet = workbook.active

	# for key,value in values_dict:
	# 	print(key)
	
	# for key,value in values_dict:
	# 	logging.debug('The key value to use is:  %s' % (key))
	# 	logging.debug('The value of the key is:  ')
	# 	logging.debug(value)
	# 	sheet[key] = value

	# for testing
	# for k,v in values_dict.items():
	# 	print(k)
	# 	print(v)

	for k,v in values_dict.items():
		logging.debug('The key value to use is:  %s' % (k))
		logging.debug('The value of the key is:  ')
		logging.debug(v)
		sheet[k] = v

# store values up to N row
# row_analyzer(values_uptoN_list,1,position_row_N + 1)

row_analyzer(value_uptoN_dict,1,position_row_N + 1)

# store values after the number of M blank rows is inserted
# this runs to the end of the remaining rows hence upper_row_max + 1 is the upper limit

# row_analyzer(values_afterM_list,position_row_N + 2,upper_row_max + 1) # we use +2 to be the row after the one we want to insert the gaps after, we're not inserting the gaps yet, we're simply splitting the data up and storing it in 2 separate lists i.e. one before N and one for after N

row_analyzer(values_afterM_dict,position_row_N + 2,upper_row_max + 1,blank_row_num_M) # we use +2 to be the row after the one we want to insert the gaps after, we're not inserting the gaps yet, we're simply splitting the data up and storing it in 2 separate lists i.e. one before N and one for after N


# for testing
# logging.debug('The values up to N list is')
# logging.debug(values_uptoN_list)

# logging.debug('The values after M blank insertion list is')
# logging.debug(values_afterM_list)

logging.debug('The values up to N list is')
logging.debug(value_uptoN_dict)

# logging.debug('The values after M blank insertion list is')
# logging.debug(values_afterM_dict)

# create new spreadsheet to store values

nwb = openpyxl.Workbook()
# nsheet = nwb.active


# write the values from values_uptoN_list

row_builder(value_uptoN_dict,nwb)

# write the values from values_afterM_list
# ensure you add blank rows M after finishing values_uptoN_list

row_builder(values_afterM_dict,nwb)

# save the final sheet

nwb.save('alteredSheet.xlsx')
logging.debug('Spreadsheet file saved.')

